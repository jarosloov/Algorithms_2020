import requests
from bs4 import BeautifulSoup
import openpyxl
import time

try:

	res = requests.get('https://pogoda1.ru/katalog/sverdlovsk-oblast/temperatura-vody/')
	soup = BeautifulSoup(res.content)
	temperature_ = []
	name_ = []

	for table in soup.select('.x-table > .x-row'):
		temperature = table.select_one('.x-cell-water-temp').get_text(strip=True)
		links = [a for a in table.select('.x-cell > .link')]
		name = links[0].text
		temperature_.append(temperature)
		name_.append(name)

	my_wb = openpyxl.Workbook()
	my_sheet = my_wb.active
	name_rek = my_sheet.cell(row = 1, column = 1)
	name_rek.value = 'Река:'
	temperature_rek = my_sheet.cell(row = 1, column = 2)
	temperature_rek.value = 'С°'
	time_ = my_sheet.cell(row=1, column=3)
	time_.value = time.ctime()
	for num in range(len(name_)):
		test = my_sheet.cell(row = num + 2, column = 1)
		new_name = name_[num].split()
		test.value = new_name[1]
		test = my_sheet.cell(row=num + 2, column= 2)
		test.value = temperature_[num]
	my_wb.save("test.xlsx")
	print('готово')
except:
	print("Ошибка")
